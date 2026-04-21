"""Download property images and create an Excel catalog.

For each hotel property:
  1. Collects all unique images from extracted content
  2. Downloads them concurrently with proper naming
  3. Creates an Excel file with: title, URL, description, local filename
  4. Stores images in: output/<job_id>/<property>/images/

Skips tiny images (icons, spacers, tracking pixels) by size and URL patterns.
"""

import asyncio
import hashlib
import logging
import mimetypes
import re
import time
from pathlib import Path
from urllib.parse import urlparse, unquote

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

logger = logging.getLogger(__name__)

# ── Skip patterns for non-useful images ──────────────────────────────────────

_SKIP_URL_PATTERNS = [
    re.compile(r"\b(icon|favicon|logo|sprite|spacer|pixel|tracker|badge|button)\b", re.I),
    re.compile(r"\b(newsletter|subscribe|tripadvisor|google-review|google-reviews|rating|review|footer|footlogo)\b", re.I),
    re.compile(r"\.(svg|gif|ico)(\?|$)", re.I),
    re.compile(r"(googletagmanager|facebook|analytics|doubleclick|adsense)", re.I),
    re.compile(r"data:image/", re.I),
    re.compile(r"1x1|2x2|1\.png|blank\.png", re.I),
]

# Minimum file size to keep (skip tiny images like tracking pixels)
_MIN_IMAGE_BYTES = 5_000  # 5KB


def _should_skip_image(url: str) -> bool:
    """Return True if this image URL looks like an icon/tracker/spacer."""
    for pattern in _SKIP_URL_PATTERNS:
        if pattern.search(url):
            return True
    return False


def _sanitize_filename_component(raw: str) -> str:
    """Turn a title/alt/url stem into a safe filename component."""
    cleaned = re.sub(r"[^\w\s-]", "", raw or "")
    cleaned = re.sub(r"[\s_]+", "-", cleaned).strip("-").lower()
    if len(cleaned) > 80:
        cleaned = cleaned[:80].rstrip("-")
    return cleaned


def _looks_like_sentence(raw: str) -> bool:
    """Return True when a title/alt reads like long SEO copy instead of a name."""
    words = re.findall(r"[a-z0-9]+", (raw or "").lower())
    return len(words) >= 10


def _looks_generic_stem(raw: str) -> bool:
    """Return True when a URL stem is too generic to be useful as a filename."""
    stem = _sanitize_filename_component(raw)
    if not stem or len(stem) < 2:
        return True
    # Pure numeric or hash-like stems
    if re.fullmatch(r"\d+", stem):
        return True
    # Single generic word optionally followed by digits
    if re.fullmatch(r"(img|image|banner|slider|photo|pic|untitled|default|placeholder|thumb|thumbnail|asset|media|file|background|bg|hero|cover|main|feature|static)[-_]?\d*", stem):
        return True
    # Generic compound names like "room-image", "hotel-photo", "property-image" etc.
    _GENERIC_NOUNS = {"image", "photo", "pic", "img", "picture", "banner", "thumb", "thumbnail"}
    parts = re.split(r"[-_]", stem)
    if parts and parts[-1] in _GENERIC_NOUNS:
        return True
    return False


def _make_safe_filename(title: str, url: str, index: int, description: str = "") -> str:
    """Create a safe, descriptive filename from the website's own asset hints.

    Returns something like: '001_deluxe-room-interior.jpg'
    """
    path = unquote(urlparse(url).path)
    raw_stem = Path(path).stem
    # Strip React/webpack build hashes: "room-image.abc123def" → "room-image"
    # Hashes appear as a dotted suffix of 6+ hex chars, e.g. ".3f7a2c1d"
    url_stem = re.sub(r"\.[0-9a-f]{6,}$", "", raw_stem, flags=re.IGNORECASE)
    title_name = _sanitize_filename_component(title)
    description_name = _sanitize_filename_component(description)
    url_name = _sanitize_filename_component(url_stem)

    # Build candidates in priority order — best descriptive name first.
    # Generic stems (room-image, hotel-photo, img001) are never used; fall
    # back to a URL hash so every image gets a unique, stable name.
    candidates: list[str] = []

    # Apply _looks_generic_stem to title/alt too — catches cases where the
    # website's alt text is literally "Room Image", "Hotel Photo", etc.
    if title_name and not _looks_like_sentence(title) and not _looks_generic_stem(title_name):
        candidates.append(title_name)
    if url_name and not _looks_generic_stem(url_stem):
        candidates.append(url_name)
    if description_name and not _looks_like_sentence(description):
        candidates.append(description_name)
    # Long-sentence titles are still better than nothing if nothing else worked —
    # but still exclude generic stems so "room-image" never sneaks in via fallback.
    if title_name and not _looks_generic_stem(title_name):
        candidates.append(title_name[:40].rstrip("-"))
    if description_name and not _looks_generic_stem(description_name):
        candidates.append(description_name[:40].rstrip("-"))

    name = next((candidate for candidate in candidates if candidate and len(candidate) >= 2), "")
    if not name:
        # No usable descriptive name — use a short URL hash for uniqueness
        name = hashlib.md5(url.encode()).hexdigest()[:10]

    # Get extension from URL
    ext = _get_extension(url)

    return f"{index:03d}_{name}{ext}"


def _get_extension(url: str) -> str:
    """Extract image extension from URL."""
    path = urlparse(url).path.lower()
    for ext in (".jpg", ".jpeg", ".png", ".webp", ".avif", ".bmp", ".tiff"):
        if path.endswith(ext):
            return ext
    # Default to .jpg
    return ".jpg"


async def _download_one_image(
    client: httpx.AsyncClient,
    url: str,
    save_path: Path,
    timeout: int = 15,
) -> bool:
    """Download a single image. Returns True on success."""
    try:
        resp = await client.get(url, timeout=timeout)
        if resp.status_code != 200:
            return False

        content = resp.content
        if len(content) < _MIN_IMAGE_BYTES:
            logger.debug("    Skipping tiny image (%d bytes): %s", len(content), url[:80])
            return False

        # Verify it's actually an image
        content_type = resp.headers.get("content-type", "")
        if content_type and not content_type.startswith("image/"):
            return False

        save_path.parent.mkdir(parents=True, exist_ok=True)
        save_path.write_bytes(content)
        return True

    except Exception as exc:
        logger.debug("    Image download failed %s: %s", url[:60], exc)
        return False


def _create_excel_catalog(
    images_data: list[dict],
    excel_path: Path,
    property_name: str,
) -> None:
    """Create an Excel file cataloging all downloaded images.

    Columns: #, Title, URL, Description, Source Page, Filename, File Size
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Images"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["#", "Title", "Image URL", "Description", "Source Page", "Filename", "File Size (KB)"]
    col_widths = [5, 35, 50, 45, 50, 35, 12]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    # Data rows
    wrap_align = Alignment(vertical="top", wrap_text=True)
    for row_idx, img in enumerate(images_data, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = wrap_align
        ws.cell(row=row_idx, column=2, value=img.get("title", "")).alignment = wrap_align
        ws.cell(row=row_idx, column=3, value=img.get("url", "")).alignment = wrap_align
        ws.cell(row=row_idx, column=4, value=img.get("description", "")).alignment = wrap_align
        ws.cell(row=row_idx, column=5, value=img.get("source_page", "")).alignment = wrap_align
        ws.cell(row=row_idx, column=6, value=img.get("filename", "")).alignment = wrap_align
        ws.cell(row=row_idx, column=7, value=img.get("size_kb", 0)).alignment = wrap_align

    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.cell(row=1, column=1, value="Property").font = Font(bold=True)
    ws_summary.cell(row=1, column=2, value=property_name)
    ws_summary.cell(row=2, column=1, value="Total Images").font = Font(bold=True)
    ws_summary.cell(row=2, column=2, value=len(images_data))
    ws_summary.cell(row=3, column=1, value="Total Size (MB)").font = Font(bold=True)
    total_kb = sum(img.get("size_kb", 0) for img in images_data)
    ws_summary.cell(row=3, column=2, value=round(total_kb / 1024, 2))
    ws_summary.column_dimensions["A"].width = 18
    ws_summary.column_dimensions["B"].width = 40

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(excel_path))
    logger.info("  [IMAGES] Excel catalog saved: %s (%d images)", excel_path.name, len(images_data))


async def download_property_images(
    property_name: str,
    extracted_pages: list[dict],
    output_dir: Path,
    max_concurrent: int = 10,
) -> dict:
    """Download all images for a property and create an Excel catalog.

    Args:
        property_name: Display name of the property.
        extracted_pages: List of extracted content dicts (from content_extractor).
        output_dir: Base output directory (e.g., output/<job_id>/<property>/).
        max_concurrent: Max concurrent downloads.

    Returns:
        dict with: total_found, downloaded, skipped, failed, excel_path, images_dir
    """
    t0 = time.perf_counter()
    logger.info(">>> [IMAGES] Starting image download for property: %s", property_name)

    # Collect all unique images across all pages
    all_images: list[dict] = []
    seen_urls: set[str] = set()

    for page in extracted_pages:
        for img in page.get("images", []):
            url = img.get("src", "")
            if not url or url in seen_urls:
                continue
            if _should_skip_image(url):
                continue
            seen_urls.add(url)
            all_images.append({
                "url": url,
                "title": img.get("title", "") or img.get("alt", ""),
                "alt": img.get("alt", ""),
                "description": img.get("description", ""),
                "source_page": img.get("source_page", page.get("source_url", "")),
            })

    total_found = len(all_images)
    logger.info("  [IMAGES] Found %d unique images for %s", total_found, property_name)

    if total_found == 0:
        return {
            "total_found": 0, "downloaded": 0, "skipped": 0, "failed": 0,
            "excel_path": None, "images_dir": None,
        }

    # Create images directory
    images_dir = output_dir / "images"
    images_dir.mkdir(parents=True, exist_ok=True)

    # Download images concurrently
    semaphore = asyncio.Semaphore(max_concurrent)
    downloaded_data: list[dict] = []
    failed = 0
    skipped = 0

    async def _download_with_semaphore(img: dict, index: int):
        nonlocal failed, skipped
        async with semaphore:
            url = img["url"]
            filename = _make_safe_filename(img["title"], url, index, img["description"])
            save_path = images_dir / filename

            async with httpx.AsyncClient(
                follow_redirects=True,
                headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/124.0.0.0"},
            ) as client:
                success = await _download_one_image(client, url, save_path)

            if success:
                size_kb = round(save_path.stat().st_size / 1024, 1)
                downloaded_data.append({
                    "title": img["title"],
                    "url": url,
                    "description": img["description"],
                    "source_page": img["source_page"],
                    "filename": filename,
                    "size_kb": size_kb,
                    "local_path": str(save_path),
                })
            else:
                if save_path.exists():
                    save_path.unlink()  # Clean up tiny/invalid files
                    skipped += 1
                else:
                    failed += 1

    # Run all downloads
    tasks = [_download_with_semaphore(img, i + 1) for i, img in enumerate(all_images)]
    await asyncio.gather(*tasks)

    # Sort by filename for consistent ordering
    downloaded_data.sort(key=lambda x: x["filename"])

    # Re-number after sorting
    for i, img in enumerate(downloaded_data):
        old_name = img["filename"]
        new_name = f"{i + 1:03d}_{old_name.split('_', 1)[1]}" if "_" in old_name else old_name
        if old_name != new_name:
            old_path = images_dir / old_name
            new_path = images_dir / new_name
            if old_path.exists():
                old_path.rename(new_path)
                img["filename"] = new_name
                img["local_path"] = str(new_path)

    # Create Excel catalog
    safe_prop = re.sub(r'[^\w\s-]', '', property_name).strip().replace(' ', '_').lower()
    excel_path = output_dir / f"{safe_prop}_images.xlsx"
    if downloaded_data:
        _create_excel_catalog(downloaded_data, excel_path, property_name)
    else:
        excel_path = None

    elapsed = time.perf_counter() - t0
    logger.info(
        "<<< [IMAGES] %s: %d downloaded, %d skipped (tiny), %d failed out of %d found (%.1fs)",
        property_name, len(downloaded_data), skipped, failed, total_found, elapsed,
    )

    return {
        "total_found": total_found,
        "downloaded": len(downloaded_data),
        "skipped": skipped,
        "failed": failed,
        "excel_path": str(excel_path) if excel_path else None,
        "images_dir": str(images_dir),
    }
