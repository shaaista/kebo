from __future__ import annotations

import argparse
import asyncio
import os
from pathlib import Path
from typing import Any
import sys

from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

DEFAULT_XLSX_PATH = Path(r"C:\Users\Hp\Downloads\main_images 1.xlsx")
DEFAULT_HOTEL_CODE = "iconiqa_test2"


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("-", "_").replace(" ", "_")
    while "__" in text:
        text = text.replace("__", "_")
    return text.strip("_")


def _clean_cell(value: Any) -> str:
    text = str(value or "").strip()
    if text.lower() in {"none", "nan", "null", "n/a", "na"}:
        return ""
    return text


def _resolve_column(headers: list[str], candidates: list[str]) -> int:
    normalized = {_normalize_header(name): idx for idx, name in enumerate(headers)}
    for candidate in candidates:
        idx = normalized.get(_normalize_header(candidate))
        if idx is not None:
            return idx
    return -1


def _read_rows_from_xlsx(xlsx_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active

    row_iter = ws.iter_rows(min_row=1, values_only=True)
    header_row = next(row_iter, None)
    if not header_row:
        return []
    headers = [str(item or "").strip() for item in header_row]

    title_idx = _resolve_column(headers, ["title", "image_title", "label", "name"])
    url_idx = _resolve_column(headers, ["image_url", "url", "image", "image_link", "imageurl"])
    category_idx = _resolve_column(headers, ["category", "type", "label_type", "tag"])
    description_idx = _resolve_column(headers, ["description", "caption", "details", "copy"])
    filename_idx = _resolve_column(headers, ["filename", "file_name", "image_name"])
    if title_idx < 0 or url_idx < 0:
        raise ValueError("XLSX must include columns for title and image_url/url.")

    rows: list[dict[str, Any]] = []
    for values in row_iter:
        values = values or []
        title = _clean_cell(values[title_idx] if title_idx < len(values) else "")
        image_url = _clean_cell(values[url_idx] if url_idx < len(values) else "")
        category = _clean_cell(values[category_idx] if category_idx >= 0 and category_idx < len(values) else "")
        description = _clean_cell(values[description_idx] if description_idx >= 0 and description_idx < len(values) else "")
        filename = _clean_cell(values[filename_idx] if filename_idx >= 0 and filename_idx < len(values) else "")
        if not title and description:
            title = description.split(".")[0][:120].strip()
        if not title and filename:
            stem = Path(filename).stem
            title = stem.replace("_", " ").replace("-", " ").strip()
            title = " ".join(title.split())
        if not title or not image_url:
            continue
        rows.append(
            {
                "title": title,
                "image_url": image_url,
                "description": description,
                "category": category,
            }
        )
    return rows


async def _run(args) -> None:
    if args.database_url:
        os.environ["DATABASE_URL"] = str(args.database_url).strip()

    from models.database import init_db
    from services.image_tool_service import image_tool_service

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")

    await init_db()
    rows = _read_rows_from_xlsx(xlsx_path)
    if not rows:
        print("No valid rows found in XLSX.")
        return

    upsert_result = await image_tool_service.upsert_assets(
        hotel_code=args.hotel_code,
        rows=rows,
        source_label=str(xlsx_path.name),
        replace_existing=not args.append,
    )
    tool_registered = await image_tool_service.ensure_tool_registered(
        hotel_code=args.hotel_code,
        enabled=not args.disable_tool,
    )

    print(f"Hotel code         : {args.hotel_code}")
    print(f"XLSX source        : {xlsx_path}")
    print(f"Rows parsed        : {len(rows)}")
    print(f"Inserted           : {upsert_result.get('inserted', 0)}")
    print(f"Updated            : {upsert_result.get('updated', 0)}")
    print(f"Skipped            : {upsert_result.get('skipped', 0)}")
    print(f"Replace existing   : {not args.append}")
    print(f"Tool registered    : {tool_registered}")
    print(f"Tool enabled       : {not args.disable_tool}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Import hotel image assets from XLSX and register smart image tool.")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX_PATH), help="Path to source XLSX file.")
    parser.add_argument("--hotel-code", default=DEFAULT_HOTEL_CODE, help="Target hotel code.")
    parser.add_argument("--append", action="store_true", help="Append/upsert instead of replacing existing hotel images.")
    parser.add_argument("--disable-tool", action="store_true", help="Register the tool but keep it disabled.")
    parser.add_argument("--database-url", default="", help="Optional DB URL override for this import run.")
    args = parser.parse_args()
    asyncio.run(_run(args))


if __name__ == "__main__":
    main()
